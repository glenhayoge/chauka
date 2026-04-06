"""Excel export endpoints ported from the original Django export views."""
from __future__ import annotations

import io
import re
import textwrap
from collections import defaultdict
from datetime import date, timedelta
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.auth.dependencies import get_current_user
from app.database import get_db
from app.models.appconf import Settings
from app.models.contacts import User
from app.models.logframe import (
    Activity,
    Assumption,
    BudgetLine,
    DataEntry,
    Indicator,
    Logframe,
    Period,
    Result,
    RiskRating,
    SubIndicator,
    Target,
)
from app.models.org import Organisation, Program, Project
from app.services.resolve import resolve_logframe

router = APIRouter(
    prefix="/api/logframes/{logframe_public_id}/export",
    tags=["export"],
)

# ---------------------------------------------------------------------------
# Styling constants (ported from Django export/views.py)
# ---------------------------------------------------------------------------
GANTT_COL_WIDTH = 4
DATE_COL_WIDTH = 11
MAX_WIDTH = 60

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _solid(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


GANTT_FILL = _solid("8FBC8F")

_DEFAULT_STYLE: dict = {}
LEVEL_STYLES: dict = defaultdict(lambda: _DEFAULT_STYLE)
LEVEL_STYLES[1] = {"font": Font(color="FFFFFF"), "fill": _solid("505A6B")}
LEVEL_STYLES[2] = LEVEL_STYLES[1]
LEVEL_STYLES[3] = LEVEL_STYLES[2]
LEVEL_STYLES[5] = {"fill": _solid("FFFF00")}
LEVEL_STYLES["activity"] = {"fill": _solid("FFFFCC")}
LEVEL_STYLES["indicator"] = {"fill": _solid("FFF47F")}
LEVEL_STYLES["subindicator-header"] = {"fill": _solid("99CC99")}
LEVEL_STYLES["subindicator-name"] = {"fill": _solid("FFFFFF")}
LEVEL_STYLES["subindicator-milestone"] = {"fill": _solid("FFFFFF")}
LEVEL_STYLES["subindicator-total"] = {"fill": _solid("D0E1F4")}
LEVEL_STYLES["subindicator-rating"] = {"fill": _solid("FFCC99")}


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------
_TAG_RE = re.compile(r"<[^>]+>")


def _html2txt(html: str | None) -> str:
    if not html:
        return ""
    return _TAG_RE.sub("", html).strip()


def _apply_styles(cell, styles: dict):
    for key, value in styles.items():
        setattr(cell, key, value)


def _styled_cell(value, styles: dict | None = None) -> dict:
    d: dict = {"value": value}
    if styles:
        d["styles"] = styles
    return d


def _row_style(styles: dict, cells: list) -> list[dict]:
    return [_styled_cell(c, styles) for c in cells]


# ---------------------------------------------------------------------------
# Period utilities (ported from logframe/period_utils.py)
# ---------------------------------------------------------------------------

def _get_month_shift(month: int, num_periods: int, period_offset: int = 1):
    new_month = month + (12 // num_periods) * period_offset
    add_year = 0 if new_month < 13 else 1
    new_month = new_month if new_month < 13 else new_month % 12
    if new_month == 0:
        new_month = 12
    return new_month, add_year


def _get_period(start_date: date, num_periods: int):
    new_month, add_year = _get_month_shift(start_date.month, num_periods)
    next_period = date(start_date.year + add_year, new_month, 1)
    return start_date, next_period - timedelta(days=1)


def _get_periods(start_date: date, end_date: date, year_start: int, num_periods: int):
    periods_begin = [_get_month_shift(year_start, num_periods, p) for p in range(num_periods)]
    periods = []
    for year in range(start_date.year - 1, end_date.year + 1):
        for month, year_carry in periods_begin:
            periods.append(date(year + year_carry, month, 1))
    start_idx = end_idx = 0
    for i, period in enumerate(periods):
        if period <= start_date:
            start_idx = i
        if period > end_date:
            end_idx = i
            break
    if not end_idx:
        end_idx = len(periods)
    return periods[start_idx:end_idx]


def _periods_intersect(s, e, x, y) -> bool:
    if not y and not x:
        return True
    if not y:
        return x <= e
    if not x:
        return y >= s
    return (x >= s and x <= e) or (x <= s and y >= s)


def _get_period_list(start: date, end: date):
    starts = _get_periods(start, end, start.month, 12)
    return [_get_period(s, 12) for s in starts]


def _get_period_header(start: date, end: date | None = None, padding: int = 0) -> list[str]:
    months = MONTHS[start.month - 1:] + MONTHS[:start.month - 1]
    if end:
        length = end.month - start.month + 1
        if length <= 0:
            length += 12
        months = months[:length]
    return [""] * padding + months


def _mark_row(start_date, end_date, periods, padding: int = 0) -> list:
    marked = []
    marked_value = _styled_cell("", {"fill": GANTT_FILL})
    for ps, pe in periods:
        marked.append(marked_value if _periods_intersect(ps, pe, start_date, end_date) else "")
    return [""] * padding + marked


# ---------------------------------------------------------------------------
# XLSX generation (simplified from SpreadsheetResponseMixin)
# ---------------------------------------------------------------------------

def _generate_xlsx(data: list[list], format_ws=None) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active

    max_widths: dict[int, int] = defaultdict(int)

    for r, row in enumerate(data, 1):
        row_heights = [-1]
        for c, cellval in enumerate(row, 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cellval, dict):
                value = cellval.get("value", "")
                if "styles" in cellval:
                    _apply_styles(cell, cellval["styles"])
            else:
                value = cellval

            if isinstance(value, str):
                vlen = len(value)
                if vlen > max_widths[c]:
                    max_widths[c] = vlen
                if vlen > MAX_WIDTH:
                    cell.alignment = Alignment(wrap_text=True)
                    row_heights.append(len(textwrap.wrap(value, MAX_WIDTH)))

            cell.value = value

        if row:
            h = max(row_heights)
            if h > 1:
                ws.row_dimensions[r].height = h * 15

    for c, w in max_widths.items():
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = min(w + 2, MAX_WIDTH)

    if format_ws:
        format_ws(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_response(buf: io.BytesIO, filename: str) -> StreamingResponse:
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Data loading helpers
# ---------------------------------------------------------------------------

async def _load_settings(db: AsyncSession, logframe_id: int) -> Settings:
    result = await db.execute(
        select(Settings).where(Settings.logframe_id == logframe_id)
    )
    settings = result.scalar_one_or_none()
    if not settings:
        raise HTTPException(404, "Logframe settings not found")
    return settings


async def _load_results_tree(db: AsyncSession, logframe_id: int) -> list[Result]:
    """Load results in parent-child order (pre-order traversal)."""
    stmt = (
        select(Result)
        .where(Result.logframe_id == logframe_id)
        .options(
            selectinload(Result.indicators).selectinload(Indicator.subindicators),
            selectinload(Result.activities),
        )
        .order_by(Result.order)
    )
    result = await db.execute(stmt)
    all_results = result.scalars().all()

    children_map: dict[int | None, list[Result]] = defaultdict(list)
    for r in all_results:
        children_map[r.parent_id].append(r)

    ordered: list[Result] = []

    def walk(parent_id: int | None):
        for child in children_map.get(parent_id, []):
            ordered.append(child)
            walk(child.id)

    walk(None)
    return ordered


async def _load_targets_for_period(db: AsyncSession, period_id: int) -> dict[int, str]:
    """Map subindicator_id → target value for a given period (milestone)."""
    stmt = select(Target).where(Target.milestone_id == period_id)
    result = await db.execute(stmt)
    return {t.subindicator_id: t.value or "" for t in result.scalars().all()}


async def _load_latest_data_entries(db: AsyncSession, logframe_id: int) -> dict[int, str]:
    """Map subindicator_id → most recent data entry value."""
    stmt = (
        select(DataEntry)
        .join(DataEntry.column)
        .where(DataEntry.column.has(logframe_id=logframe_id))
    )
    result = await db.execute(stmt)
    entries: dict[int, str] = {}
    for e in result.scalars().all():
        entries[e.subindicator_id] = e.data or ""
    return entries


# ---------------------------------------------------------------------------
# Quarterly Report Export
# ---------------------------------------------------------------------------

@router.get("/quarterly-report")
async def export_quarterly_report(
    logframe_public_id: UUID,
    period: str = Query(..., description="Period as MM-YYYY"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    logframe_id = (await resolve_logframe(logframe_public_id, db)).id
    settings = await _load_settings(db, logframe_id)

    # Parse period parameter
    try:
        parts = period.split("-")
        month = int(parts[0])
        year = int(parts[1])
    except (ValueError, IndexError):
        raise HTTPException(400, "Period must be MM-YYYY format")

    # Find the matching period
    stmt = select(Period).where(
        Period.logframe_id == logframe_id,
        Period.start_month == month,
        Period.start_year == year,
    )
    result = await db.execute(stmt)
    period_obj = result.scalar_one_or_none()

    # Calculate period boundaries
    period_start = date(year, month, 1)
    if period_obj:
        period_end = date(period_obj.end_year, period_obj.end_month, 1)
        period_end = date(period_end.year, period_end.month + 1, 1) - timedelta(days=1) \
            if period_end.month < 12 else date(period_end.year, 12, 31)
    else:
        n = settings.n_periods if settings.n_periods else 4
        _, period_end = _get_period(period_start, n)

    # Load data
    results = await _load_results_tree(db, logframe_id)
    targets = await _load_targets_for_period(db, period_obj.id if period_obj else 0)
    latest_data = await _load_latest_data_entries(db, logframe_id)

    # Build spreadsheet data
    data: list[list] = []
    data.append(["Quarterly report", "", period_start.isoformat(), period_end.isoformat()])
    data.append([])

    children_map: dict[int | None, list[Result]] = defaultdict(list)
    for r in results:
        children_map[r.parent_id].append(r)

    for res in results:
        row = _row_style(
            LEVEL_STYLES[res.level] if res.level else _DEFAULT_STYLE,
            [res.name or "", _html2txt(res.description), "", ""],
        )
        data.append([""] + row)

        for indicator in sorted(res.indicators, key=lambda i: i.order):
            ind_row = _row_style(
                LEVEL_STYLES["indicator"],
                [indicator.name or "", _html2txt(indicator.description), "", ""],
            )
            data.append([""] + ind_row)

            header = _row_style(
                LEVEL_STYLES["subindicator-header"],
                ["", "Milestone for period", "Total to date", "Rating"],
            )
            data.append([""] + header)

            for sub in sorted(indicator.subindicators, key=lambda s: s.order):
                sub_row = [
                    _styled_cell(sub.name or "", LEVEL_STYLES["subindicator-name"]),
                    _styled_cell(targets.get(sub.id, ""), LEVEL_STYLES["subindicator-milestone"]),
                    _styled_cell(latest_data.get(sub.id, ""), LEVEL_STYLES["subindicator-total"]),
                    _styled_cell("", LEVEL_STYLES["subindicator-rating"]),
                ]
                data.append([""] + sub_row)
            data.append([])
        data.append([])

    buf = _generate_xlsx(data)
    return _xlsx_response(buf, f"quarterly_report_{period}.xlsx")


# ---------------------------------------------------------------------------
# Annual Plan Export
# ---------------------------------------------------------------------------

@router.get("/annual-plan")
async def export_annual_plan(
    logframe_public_id: UUID,
    year: int = Query(..., description="Year as YYYY"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    logframe_id = (await resolve_logframe(logframe_public_id, db)).id
    settings = await _load_settings(db, logframe_id)
    start_date = date(year, settings.start_month, 1)
    year_end = date(year + 1, settings.start_month, 1) - timedelta(days=1)

    results = await _load_results_tree(db, logframe_id)
    children_map: dict[int | None, list[Result]] = defaultdict(list)
    for r in results:
        children_map[r.parent_id].append(r)

    periods = _get_period_list(*_get_period(start_date, 1))
    month_headers = _get_period_header(start_date)

    data: list[list] = []
    data.append([f"{year} Annual Report"])
    data.append(["Name", "Description", "Deliverables", "Start", "End"] + month_headers)
    data.append([])

    for res in results:
        is_leaf = len(children_map.get(res.id, [])) == 0
        style = LEVEL_STYLES[res.level] if res.level else _DEFAULT_STYLE
        data.append(_row_style(style, [_html2txt(res.name), _html2txt(res.description)]))

        if is_leaf:
            activities = sorted(res.activities, key=lambda a: (a.start_date or date.max, a.order))
            for act in activities:
                if act.start_date and act.end_date and act.start_date > act.end_date:
                    continue
                act_row = _row_style(
                    LEVEL_STYLES["activity"],
                    [
                        _html2txt(act.name),
                        _html2txt(act.description),
                        _html2txt(act.deliverables),
                        act.start_date.isoformat() if act.start_date else "",
                        act.end_date.isoformat() if act.end_date else "",
                    ],
                )
                act_row += _mark_row(act.start_date, act.end_date, periods)
                data.append(act_row)
            data.append([])

    def format_ws(ws):
        for col in "FGHIJKLMNOPQ":
            ws.column_dimensions[col].width = GANTT_COL_WIDTH
        ws.column_dimensions["D"].width = DATE_COL_WIDTH
        ws.column_dimensions["E"].width = DATE_COL_WIDTH

    buf = _generate_xlsx(data, format_ws)
    return _xlsx_response(buf, f"{year}_annual_plan.xlsx")


# ---------------------------------------------------------------------------
# Quarterly Plan Export
# ---------------------------------------------------------------------------

@router.get("/quarterly-plan")
async def export_quarterly_plan(
    logframe_public_id: UUID,
    period: str = Query(..., description="Period as MM-YYYY"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    logframe_id = (await resolve_logframe(logframe_public_id, db)).id
    settings = await _load_settings(db, logframe_id)

    try:
        parts = period.split("-")
        month = int(parts[0])
        year = int(parts[1])
    except (ValueError, IndexError):
        raise HTTPException(400, "Period must be MM-YYYY format")

    n_periods = settings.n_periods if settings.n_periods else 4
    start_date, end_date = _get_period(date(year, month, 1), n_periods)
    period_length = 12 // n_periods

    results = await _load_results_tree(db, logframe_id)
    children_map: dict[int | None, list[Result]] = defaultdict(list)
    for r in results:
        children_map[r.parent_id].append(r)

    periods = _get_period_list(start_date, end_date)
    month_headers = _get_period_header(start_date, end_date)

    title = f"{start_date.strftime('%B')} - {end_date.strftime('%B')} {start_date.year} Quarter Report"

    data: list[list] = []
    data.append([title])
    data.append(["Name", "Description", "Deliverables", "Start", "End"] + month_headers)
    data.append([])

    for res in results:
        is_leaf = len(children_map.get(res.id, [])) == 0
        style = LEVEL_STYLES[res.level] if res.level else _DEFAULT_STYLE
        data.append(_row_style(style, [_html2txt(res.name), _html2txt(res.description)]))

        if is_leaf:
            activities = sorted(res.activities, key=lambda a: (a.start_date or date.max, a.order))
            for act in activities:
                if act.start_date and act.end_date and act.start_date > act.end_date:
                    continue
                act_row = _row_style(
                    LEVEL_STYLES["activity"],
                    [
                        _html2txt(act.name),
                        _html2txt(act.description),
                        _html2txt(act.deliverables),
                        act.start_date.isoformat() if act.start_date else "",
                        act.end_date.isoformat() if act.end_date else "",
                    ],
                )
                act_row += _mark_row(act.start_date, act.end_date, periods)
                data.append(act_row)
            data.append([])

    def format_ws(ws):
        cols = "FGHIJKLMNOPQ"
        for col in cols[:period_length]:
            ws.column_dimensions[col].width = GANTT_COL_WIDTH
        ws.column_dimensions["D"].width = DATE_COL_WIDTH
        ws.column_dimensions["E"].width = DATE_COL_WIDTH

    short_start = start_date.strftime("%b").lower()
    short_end = end_date.strftime("%b").lower()
    filename = f"{short_start}-{short_end}_{year}_quarter_plan.xlsx"

    buf = _generate_xlsx(data, format_ws)
    return _xlsx_response(buf, filename)


# ---------------------------------------------------------------------------
# Full Logframe Export
# ---------------------------------------------------------------------------

@router.get("/logframe")
async def export_logframe(
    logframe_public_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    logframe_id = (await resolve_logframe(logframe_public_id, db)).id
    """Export the full logframe (results, indicators, activities, assumptions) to Excel."""
    # Load logframe name
    lf_result = await db.execute(select(Logframe).where(Logframe.id == logframe_id))
    logframe = lf_result.scalar_one_or_none()
    if not logframe:
        raise HTTPException(404, "Logframe not found")

    settings = await db.execute(select(Settings).where(Settings.logframe_id == logframe_id))
    settings = settings.scalar_one_or_none()

    currency = settings.currency if settings else ""
    start_year = settings.start_year if settings else ""
    end_year = settings.end_year if settings else ""
    period_str = str(start_year) if start_year == end_year else f"{start_year} – {end_year}"

    # Load results tree with all relationships
    stmt = (
        select(Result)
        .where(Result.logframe_id == logframe_id)
        .options(
            selectinload(Result.indicators).selectinload(Indicator.subindicators),
            selectinload(Result.activities).selectinload(Activity.budget_lines),
            selectinload(Result.assumptions).selectinload(Assumption.risk_rating),
            selectinload(Result.rating),
        )
        .order_by(Result.order)
    )
    res_result = await db.execute(stmt)
    all_results = res_result.scalars().all()

    # Build tree in pre-order
    children_map: dict[int | None, list[Result]] = defaultdict(list)
    for r in all_results:
        children_map[r.parent_id].append(r)

    ordered: list[Result] = []

    def walk(parent_id):
        for child in children_map.get(parent_id, []):
            ordered.append(child)
            walk(child.id)

    walk(None)

    # Load targets for all indicators
    all_indicator_ids = [i.id for r in ordered for i in r.indicators]
    targets_by_indicator: dict[int, list[Target]] = defaultdict(list)
    if all_indicator_ids:
        t_stmt = select(Target).where(Target.indicator_id.in_(all_indicator_ids))
        t_result = await db.execute(t_stmt)
        for t in t_result.scalars().all():
            targets_by_indicator[t.indicator_id].append(t)

    # Level labels (1=Impact, 2=Outcome, 3=Output, etc.)
    level_labels: dict[int, str] = {}
    if settings and settings.level_labels:
        level_labels = {int(k): v for k, v in settings.level_labels.items()}
    default_labels = {1: "Impact", 2: "Outcome", 3: "Output", 4: "Activity"}

    def level_label(level: int | None) -> str:
        if level is None:
            return ""
        return level_labels.get(level) or default_labels.get(level) or f"Level {level}"

    # ---------------------------------------------------------------------------
    # Build spreadsheet
    # ---------------------------------------------------------------------------
    HEADER_STYLE = {"font": Font(bold=True, color="FFFFFF"), "fill": _solid("374151")}
    LEVEL1_STYLE = {"font": Font(bold=True, color="FFFFFF"), "fill": _solid("1e3a5f")}
    LEVEL2_STYLE = {"font": Font(bold=True, color="FFFFFF"), "fill": _solid("374151")}
    LEVEL3_STYLE = {"font": Font(bold=True), "fill": _solid("e5e7eb")}
    INDICATOR_STYLE = {"fill": _solid("fefce8")}
    ACTIVITY_STYLE = {"fill": _solid("f0fdf4")}
    ASSUMPTION_STYLE = {"fill": _solid("fff7ed")}

    data: list[list] = []

    # Title
    data.append([_styled_cell(logframe.name, {"font": Font(bold=True, size=14)})])
    meta_parts = []
    if settings:
        meta_parts.append(f"Period: {period_str}")
        meta_parts.append(f"Currency: {currency}")
    data.append([", ".join(meta_parts)] if meta_parts else [])
    data.append([])

    # Column headers
    data.append(_row_style(HEADER_STYLE, [
        "Level", "Name", "Description",
        "Source of Verification", "Baseline", "Targets",
        "Start Date", "End Date", "Budget", "Risk Rating",
    ]))

    for result in ordered:
        depth = result.level or 1
        if depth == 1:
            style = LEVEL1_STYLE
        elif depth == 2:
            style = LEVEL2_STYLE
        else:
            style = LEVEL3_STYLE

        rating_name = result.rating.name if result.rating else ""
        data.append(_row_style(style, [
            level_label(result.level),
            _html2txt(result.name),
            _html2txt(result.description),
            "", "", "", "", "", "", rating_name,
        ]))

        # Indicators
        for ind in sorted(result.indicators, key=lambda i: i.order):
            subs = sorted(ind.subindicators, key=lambda s: s.order)
            targets = targets_by_indicator.get(ind.id, [])
            target_parts = []
            for t in targets:
                if t.value:
                    sub = next((s for s in subs if s.id == t.subindicator_id), None)
                    label = sub.name if sub else ""
                    target_parts.append(f"{label}: {t.value}" if label else t.value)
            target_str = "; ".join(target_parts) if target_parts else ""
            components = ", ".join(s.name for s in subs) if subs else ""
            ind_name = _html2txt(ind.name)
            if components:
                ind_name += f" [{components}]"
            data.append(_row_style(INDICATOR_STYLE, [
                "Indicator",
                ind_name,
                _html2txt(ind.description),
                ind.source_of_verification or "",
                "Required" if ind.needs_baseline else "",
                target_str,
                "", "", "", "",
            ]))

        # Activities
        for act in sorted(result.activities, key=lambda a: a.order):
            total_budget = sum(bl.amount or 0 for bl in act.budget_lines)
            budget_str = (
                f"{currency} {total_budget:,.2f}" if total_budget else ""
            )
            data.append(_row_style(ACTIVITY_STYLE, [
                "Activity",
                _html2txt(act.name),
                _html2txt(act.description),
                "", "", "",
                act.start_date.isoformat() if act.start_date else "",
                act.end_date.isoformat() if act.end_date else "",
                budget_str,
                "",
            ]))

        # Assumptions
        for assumption in result.assumptions:
            risk_name = assumption.risk_rating.name if assumption.risk_rating else ""
            data.append(_row_style(ASSUMPTION_STYLE, [
                "Assumption",
                _html2txt(assumption.description),
                "", "", "", "", "", "", "", risk_name,
            ]))

    def format_ws(ws):
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 16
        ws.column_dimensions["J"].width = 14

    safe_name = re.sub(r"[^\w\s-]", "", logframe.name)[:40].strip().replace(" ", "_")
    buf = _generate_xlsx(data, format_ws)
    return _xlsx_response(buf, f"{safe_name}_logframe.xlsx")


# ---------------------------------------------------------------------------
# Professional Multi-Format Logframe Export
# ---------------------------------------------------------------------------

@router.get("/logframe-pro")
async def export_logframe_professional(
    logframe_public_id: UUID,
    style: str = Query("donor", description="Export style: donor|expanded|simple|eu|dfat"),
    include_activities: bool = Query(True),
    include_indicators: bool = Query(True),
    include_budget: bool = Query(True),
    include_summary: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    logframe_id = (await resolve_logframe(logframe_public_id, db)).id
    """Generate a professional, donor-ready logframe Excel export."""
    from app.services.excel_export import build_logframe_excel

    # Load logframe
    lf_res = await db.execute(select(Logframe).where(Logframe.id == logframe_id))
    logframe = lf_res.scalar_one_or_none()
    if not logframe:
        raise HTTPException(404, "Logframe not found")

    # Load settings
    cfg_res = await db.execute(select(Settings).where(Settings.logframe_id == logframe_id))
    settings = cfg_res.scalar_one_or_none()

    currency   = settings.currency   if settings else "USD"
    start_year = settings.start_year if settings else None
    end_year   = settings.end_year   if settings else None

    # Level labels
    level_labels: dict[int, str] = {1: "Impact", 2: "Outcome", 3: "Output", 4: "Activity"}
    if settings and settings.level_labels:
        for k, v in settings.level_labels.items():
            try:
                level_labels[int(k)] = str(v)
            except (ValueError, TypeError):
                pass

    # Org context
    org_name = ""
    project_name = ""
    if logframe.project_id:
        proj_res = await db.execute(
            select(Project)
            .where(Project.id == logframe.project_id)
            .options(selectinload(Project.program).selectinload(Program.organisation))
        )
        proj = proj_res.scalar_one_or_none()
        if proj:
            project_name = proj.name
            if proj.program and proj.program.organisation:
                org_name = proj.program.organisation.name
    elif logframe.program_id:
        prog_res = await db.execute(
            select(Program)
            .where(Program.id == logframe.program_id)
            .options(selectinload(Program.organisation))
        )
        prog = prog_res.scalar_one_or_none()
        if prog:
            if prog.organisation:
                org_name = prog.organisation.name

    # Load results with all relationships
    stmt = (
        select(Result)
        .where(Result.logframe_id == logframe_id)
        .options(
            selectinload(Result.indicators).selectinload(Indicator.subindicators),
            selectinload(Result.activities).selectinload(Activity.budget_lines),
            selectinload(Result.assumptions).selectinload(Assumption.risk_rating),
            selectinload(Result.rating),
        )
        .order_by(Result.order)
    )
    results = list((await db.execute(stmt)).scalars().all())

    # Load targets indexed by indicator_id
    ind_ids = [i.id for r in results for i in r.indicators]
    all_targets: dict[int, list] = defaultdict(list)
    if ind_ids:
        t_res = await db.execute(select(Target).where(Target.indicator_id.in_(ind_ids)))
        for t in t_res.scalars().all():
            all_targets[t.indicator_id].append(t)

    buf = build_logframe_excel(
        logframe_name=logframe.name,
        org_name=org_name,
        project_name=project_name,
        currency=currency,
        start_year=start_year,
        end_year=end_year,
        results=results,
        all_targets=dict(all_targets),
        level_labels=level_labels,
        style=style,
        include_activities=include_activities,
        include_indicators=include_indicators,
        include_budget=include_budget,
        include_summary=include_summary,
    )

    safe_name = re.sub(r"[^\w\s-]", "", logframe.name)[:40].strip().replace(" ", "_")
    return _xlsx_response(buf, f"{safe_name}_{style}.xlsx")
