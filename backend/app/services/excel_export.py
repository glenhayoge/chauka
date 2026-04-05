"""
Chauka MEL – Professional Excel export service.
Generates donor-ready logframe reports in multiple styles:
  donor    – Full NGO donor-grade with multi-sheet workbook (default)
  expanded – Indicators as child rows beneath each result
  simple   – Compact single row per result
  eu       – EU Logical Framework (Intervention Logic format)
  dfat     – Australian DFAT Performance Assessment Framework
"""
from __future__ import annotations

import io
import re
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
VERSION = "1.0"
SYSTEM_NAME = "Chauka MEL System"

# Color palette (no leading #)
C = {
    "navy":      "1B4F72",
    "d_green":   "1E6B3C",
    "m_green":   "52BE80",
    "l_green":   "D5F5E3",
    "white":     "FFFFFF",
    "l_grey":    "F2F2F2",
    "m_grey":    "9CA3AF",
    "d_grey":    "374151",
    "l_amber":   "FEF9E7",
    "l_blue":    "EBF5FB",
    "footer":    "F1F5F9",
    "border":    "D0D0D0",
    "dark":      "1F2937",
    "alt":       "FAFAFA",
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


_T = Side(style="thin", color=C["border"])
_M = Side(style="medium", color="AAAAAA")
THIN_B  = Border(left=_T,  right=_T,  top=_T,  bottom=_T)
OUTER_B = Border(left=_M,  right=_M,  top=_M,  bottom=_M)
BOT_B   = Border(bottom=_T)

AW   = Alignment(wrap_text=True, vertical="top")
AWC  = Alignment(wrap_text=True, vertical="top", horizontal="center")
AWR  = Alignment(wrap_text=True, vertical="top", horizontal="right")
CEN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

HDR_FONT  = Font(bold=True, size=9, color=C["white"])
HDR_FILL  = _fill(C["d_grey"])

# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------
_TAG_RE = re.compile(r"<[^>]+>")


def _strip(text: str | None) -> str:
    if not text:
        return ""
    return _TAG_RE.sub("", text).strip()


def _fmt_currency(amount: float, currency: str) -> str:
    if not amount:
        return "—"
    return f"{currency} {amount:,.2f}"


def _fmt_date(d) -> str:
    if not d:
        return "—"
    try:
        return d.strftime("%d %b %Y")
    except AttributeError:
        return str(d)


def _fmt_period(start_year: int | None, end_year: int | None) -> str:
    if not start_year:
        return "—"
    if start_year == end_year:
        return str(start_year)
    return f"{start_year} – {end_year}"


def _letter(col: int) -> str:
    return get_column_letter(col)


# ---------------------------------------------------------------------------
# Hierarchy numbering
# ---------------------------------------------------------------------------

def _generate_codes(results: list) -> dict[int, str]:
    """Generate hierarchical codes like 1, 1.1, 1.1.1 for each result."""
    children_map: dict[int | None, list] = defaultdict(list)
    for r in results:
        children_map[r.parent_id].append(r)
    for lst in children_map.values():
        lst.sort(key=lambda x: x.order)

    codes: dict[int, str] = {}

    def _walk(parent_id: int | None, prefix: str) -> None:
        for i, r in enumerate(children_map.get(parent_id, []), 1):
            code = f"{prefix}{i}" if prefix else str(i)
            codes[r.id] = code
            _walk(r.id, f"{code}.")

    _walk(None, "")
    return codes


# ---------------------------------------------------------------------------
# Target formatting
# ---------------------------------------------------------------------------

def _format_targets(targets: list, subindicators: list) -> str:
    """Format targets, detecting sex-disaggregated data and combining totals."""
    if not targets:
        return "—"

    sub_map = {s.id: s.name for s in subindicators}
    female_total: float | None = None
    male_total: float | None = None
    other_parts: list[str] = []
    f_seen: set[int] = set()
    m_seen: set[int] = set()

    for t in targets:
        if not t.value:
            continue
        sub_name = sub_map.get(t.subindicator_id, "")
        sl = sub_name.lower()
        is_f = any(kw in sl for kw in ["female", "women", "girl", "# of f", "# f"])
        is_m = any(kw in sl for kw in ["male", "men", "boy", "# of m", "# m"])

        if is_f and t.subindicator_id not in f_seen:
            try:
                female_total = (female_total or 0) + float(t.value.replace(",", ""))
                f_seen.add(t.subindicator_id)
            except ValueError:
                other_parts.append(f"{sub_name}: {t.value}" if sub_name else t.value)
        elif is_m and t.subindicator_id not in m_seen:
            try:
                male_total = (male_total or 0) + float(t.value.replace(",", ""))
                m_seen.add(t.subindicator_id)
            except ValueError:
                other_parts.append(f"{sub_name}: {t.value}" if sub_name else t.value)
        else:
            lbl = f"{sub_name}: " if sub_name else ""
            other_parts.append(f"{lbl}{t.value}")

    lines: list[str] = []
    if female_total is not None or male_total is not None:
        f_n = int(female_total) if female_total is not None else 0
        m_n = int(male_total) if male_total is not None else 0
        if female_total is not None and male_total is not None:
            lines.append(f"{f_n + m_n:,} ({f_n:,}F / {m_n:,}M)")
        elif female_total is not None:
            lines.append(f"{f_n:,} (female)")
        else:
            lines.append(f"{m_n:,} (male)")

    lines.extend(other_parts)
    return "\n".join(lines) if lines else "—"


# ---------------------------------------------------------------------------
# Summary computation
# ---------------------------------------------------------------------------

def _compute_summary(results: list, currency: str) -> dict:
    level_counts: dict[int, int] = defaultdict(int)
    total_budget = 0.0
    activity_count = 0
    indicator_count = 0

    for r in results:
        if r.level:
            level_counts[r.level] += 1
        for act in r.activities:
            activity_count += 1
            for bl in act.budget_lines:
                total_budget += bl.amount or 0
        indicator_count += len(r.indicators)

    return {
        "level_counts": dict(level_counts),
        "activity_count": activity_count,
        "indicator_count": indicator_count,
        "total_budget": total_budget,
        "total_budget_fmt": _fmt_currency(total_budget, currency),
    }


# ---------------------------------------------------------------------------
# Low-level cell writers
# ---------------------------------------------------------------------------

def _cell(ws, row: int, col: int, value=None, *,
          font=None, fill=None, align=None, border=None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border


def _merge(ws, row: int, c1: int, c2: int, value=None, *,
           font=None, fill=None, align=None, border=None) -> None:
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=value)
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border


def _row_border(ws, row: int, c1: int, c2: int, border=THIN_B) -> None:
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).border = border


def _set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[_letter(i)].width = w


# ---------------------------------------------------------------------------
# Level styling
# ---------------------------------------------------------------------------

def _level_style(level: int | None) -> tuple[Font, PatternFill]:
    """Return (Font, Fill) for a result level."""
    if level == 1:
        return Font(bold=True, size=10, color=C["white"]), _fill(C["navy"])
    elif level == 2:
        return Font(bold=True, size=10, color=C["white"]), _fill(C["d_green"])
    elif level == 3:
        return Font(bold=True, size=10, color=C["dark"]),  _fill(C["l_green"])
    else:
        return Font(size=10, color=C["dark"]),              _fill(C["white"])


def _indent(depth: int) -> str:
    return "    " * depth


# ---------------------------------------------------------------------------
# Shared blocks: header + summary + footer
# ---------------------------------------------------------------------------

def _write_header(ws, start_row: int, ncols: int, *,
                  logframe_name: str, org_name: str, project_name: str,
                  period: str, currency: str, export_date: str) -> int:
    row = start_row

    # 1 – "PROJECT LOGFRAME" sub-label
    _merge(ws, row, 1, ncols, "PROJECT LOGFRAME",
           font=Font(bold=True, size=9, color="BBBBBB"),
           fill=_fill(C["navy"]), align=AWC)
    ws.row_dimensions[row].height = 16
    row += 1

    # 2 – Main title
    _merge(ws, row, 1, ncols, logframe_name,
           font=Font(bold=True, size=16, color=C["white"]),
           fill=_fill(C["navy"]), align=AWC)
    ws.row_dimensions[row].height = 38
    row += 1

    # 3 – Meta info in one merged row
    parts = []
    if org_name:      parts.append(f"Organisation: {org_name}")
    if project_name:  parts.append(f"Project: {project_name}")
    parts.append(f"Period: {period}")
    parts.append(f"Currency: {currency}")
    _merge(ws, row, 1, ncols, "   |   ".join(parts),
           font=Font(size=9, color="CCCCCC"),
           fill=_fill(C["navy"]), align=AWC)
    ws.row_dimensions[row].height = 18
    row += 1

    # 4 – Generated by + date
    _merge(ws, row, 1, ncols,
           f"Generated by {SYSTEM_NAME}   ·   Export date: {export_date}",
           font=Font(italic=True, size=8, color="AAAAAA"),
           fill=_fill(C["navy"]), align=AWC)
    ws.row_dimensions[row].height = 14
    row += 1

    # Spacer
    _merge(ws, row, 1, ncols, "", fill=_fill("F8FAFC"))
    ws.row_dimensions[row].height = 8
    row += 1

    return row


def _write_summary(ws, start_row: int, ncols: int,
                   summary: dict, level_labels: dict[int, str]) -> int:
    row = start_row

    _merge(ws, row, 1, ncols, "PROJECT SUMMARY",
           font=Font(bold=True, size=10, color=C["navy"]),
           fill=_fill(C["l_blue"]), align=AW)
    _row_border(ws, row, 1, ncols, THIN_B)
    ws.row_dimensions[row].height = 20
    row += 1

    level_text = "   |   ".join(
        f"{level_labels.get(k, f'Level {k}')}s: {v}"
        for k, v in sorted(summary["level_counts"].items())
    )
    other_text = (
        f"Activities: {summary['activity_count']}   |   "
        f"Indicators: {summary['indicator_count']}   |   "
        f"Total Budget: {summary['total_budget_fmt']}"
    )

    for line in filter(None, [level_text, other_text]):
        _merge(ws, row, 1, ncols, line,
               font=Font(size=9, color=C["dark"]),
               fill=_fill(C["l_blue"]), align=AW)
        _row_border(ws, row, 1, ncols, THIN_B)
        ws.row_dimensions[row].height = 18
        row += 1

    # Spacer
    _merge(ws, row, 1, ncols, "", fill=_fill("F8FAFC"))
    ws.row_dimensions[row].height = 8
    row += 1

    return row


def _write_footer(ws, start_row: int, ncols: int, export_date: str) -> None:
    row = start_row
    _merge(ws, row, 1, ncols, "", fill=_fill("F8FAFC"))
    ws.row_dimensions[row].height = 6
    row += 1
    _merge(ws, row, 1, ncols,
           f"{SYSTEM_NAME}   ·   Version {VERSION}   ·   Generated: {export_date}",
           font=Font(italic=True, size=8, color=C["m_grey"]),
           fill=_fill(C["footer"]), align=CEN)
    _row_border(ws, row, 1, ncols, THIN_B)
    ws.row_dimensions[row].height = 18


def _column_headers(ws, row: int, headers: list[str]) -> None:
    for c, h in enumerate(headers, 1):
        _cell(ws, row, c, h, font=HDR_FONT, fill=HDR_FILL, align=AWC, border=THIN_B)
    ws.row_dimensions[row].height = 24


# ---------------------------------------------------------------------------
# DONOR FORMAT
# ---------------------------------------------------------------------------
DONOR_HEADERS = [
    "Code", "Narrative Summary", "Indicators",
    "Baseline", "Target", "Means of Verification",
    "Timeframe", "Budget", "Assumptions / Risks",
]
DONOR_WIDTHS = [8, 48, 36, 14, 24, 28, 16, 16, 32]


def _write_donor_sheet(
    wb: Workbook, *,
    logframe_name: str, org_name: str, project_name: str,
    period: str, currency: str, export_date: str,
    results: list, children_map: dict, codes: dict[int, str],
    level_labels: dict[int, str], all_targets: dict[int, list],
    summary: dict, include_activities: bool, include_summary: bool,
) -> None:
    ws = wb.create_sheet("Logframe")
    ncols = len(DONOR_HEADERS)

    row = _write_header(
        ws, 1, ncols,
        logframe_name=logframe_name, org_name=org_name,
        project_name=project_name, period=period,
        currency=currency, export_date=export_date,
    )
    if include_summary:
        row = _write_summary(ws, row, ncols, summary, level_labels)

    _column_headers(ws, row, DONOR_HEADERS)
    freeze_row = row + 1
    row += 1

    def _write_result(r, depth: int = 0) -> None:
        nonlocal row
        code = codes.get(r.id, "")
        level = r.level or (depth + 1)
        lbl = level_labels.get(level, f"Level {level}")
        font, fill = _level_style(level)

        inds = sorted(r.indicators, key=lambda i: i.order)
        ind_text = "\n".join(f"• {_strip(i.name)}" for i in inds) or "—"
        baseline_text = "\n".join(
            "Required" if i.needs_baseline else "—" for i in inds
        ) or "—"
        target_text = "\n".join(
            _format_targets(all_targets.get(i.id, []), i.subindicators) for i in inds
        ) or "—"
        mov_text = "\n".join(
            _strip(i.source_of_verification) or "—" for i in inds
        ) or "—"

        ass_lines = []
        for a in r.assumptions:
            risk = a.risk_rating.name if a.risk_rating else ""
            desc = _strip(a.description)
            ass_lines.append(f"• {desc}" + (f"  [{risk}]" if risk else ""))
        ass_text = "\n".join(ass_lines) or "—"

        values = [
            code,
            f"{_indent(depth)}[{lbl.upper()}]  {_strip(r.name)}",
            ind_text, baseline_text, target_text, mov_text,
            "", "",  # timeframe / budget filled by activities
            ass_text,
        ]
        for c, v in enumerate(values, 1):
            _cell(ws, row, c, v, font=font, fill=fill, align=AW, border=THIN_B)
        h = max(20, min(80, (len(ind_text.split("\n")) + 1) * 14))
        ws.row_dimensions[row].height = h
        row += 1

        # Child results first
        for child in children_map.get(r.id, []):
            _write_result(child, depth + 1)

        # Activities under this result
        if include_activities:
            acts = sorted(r.activities, key=lambda a: a.order)
            for idx, act in enumerate(acts, 1):
                total_bgt = sum(bl.amount or 0 for bl in act.budget_lines)
                tf = ""
                if act.start_date and act.end_date:
                    tf = f"{_fmt_date(act.start_date)} – {_fmt_date(act.end_date)}"
                elif act.start_date:
                    tf = f"From {_fmt_date(act.start_date)}"

                act_code = f"{code}.A{idx}" if code else f"A{idx}"
                bg = _fill(C["white"]) if row % 2 == 0 else _fill(C["alt"])
                act_vals = [
                    act_code,
                    f"{_indent(depth + 1)}↳  {_strip(act.name)}",
                    "", "", "",
                    _strip(act.deliverables) or "—",
                    tf,
                    _fmt_currency(total_bgt, currency),
                    "",
                ]
                for c, v in enumerate(act_vals, 1):
                    _cell(ws, row, c, v,
                          font=Font(size=9, color=C["dark"]),
                          fill=bg, align=AW, border=THIN_B)
                ws.row_dimensions[row].height = 18
                row += 1

    for top in children_map.get(None, []):
        _write_result(top, 0)

    row += 1
    _write_footer(ws, row, ncols, export_date)
    _set_widths(ws, DONOR_WIDTHS)
    ws.freeze_panes = f"A{freeze_row}"


# ---------------------------------------------------------------------------
# EXPANDED FORMAT
# ---------------------------------------------------------------------------
EXP_HEADERS = [
    "Code", "Level", "Narrative / Indicator",
    "Baseline", "Target", "Means of Verification",
    "Start", "End", "Budget",
]
EXP_WIDTHS = [8, 12, 52, 14, 26, 30, 13, 13, 16]


def _write_expanded_sheet(
    wb: Workbook, *,
    logframe_name: str, org_name: str, project_name: str,
    period: str, currency: str, export_date: str,
    results: list, children_map: dict, codes: dict[int, str],
    level_labels: dict[int, str], all_targets: dict[int, list],
    summary: dict, include_activities: bool, include_summary: bool,
) -> None:
    ws = wb.create_sheet("Logframe")
    ncols = len(EXP_HEADERS)

    row = _write_header(
        ws, 1, ncols,
        logframe_name=logframe_name, org_name=org_name,
        project_name=project_name, period=period,
        currency=currency, export_date=export_date,
    )
    if include_summary:
        row = _write_summary(ws, row, ncols, summary, level_labels)

    _column_headers(ws, row, EXP_HEADERS)
    freeze_row = row + 1
    row += 1

    def _write_result(r, depth: int = 0) -> None:
        nonlocal row
        code = codes.get(r.id, "")
        level = r.level or (depth + 1)
        lbl = level_labels.get(level, f"Level {level}")
        font, fill = _level_style(level)

        vals = [code, lbl.upper(), f"{_indent(depth)}{_strip(r.name)}", "", "", "", "", "", ""]
        for c, v in enumerate(vals, 1):
            _cell(ws, row, c, v, font=font, fill=fill, align=AW, border=THIN_B)
        ws.row_dimensions[row].height = 22
        row += 1

        # Indicator child rows
        for ind in sorted(r.indicators, key=lambda i: i.order):
            targets = all_targets.get(ind.id, [])
            t_text = _format_targets(targets, ind.subindicators)
            ind_vals = [
                "", "INDICATOR",
                f"    {_strip(ind.name)}",
                "Required" if ind.needs_baseline else "—",
                t_text,
                _strip(ind.source_of_verification) or "—",
                "", "", "",
            ]
            for c, v in enumerate(ind_vals, 1):
                _cell(ws, row, c, v,
                      font=Font(italic=True, size=9, color="4B5563"),
                      fill=_fill(C["l_grey"]), align=AW, border=THIN_B)
            ws.row_dimensions[row].height = max(18, min(60, len(t_text.split("\n")) * 14))
            row += 1

        # Assumption child rows
        for a in r.assumptions:
            risk = a.risk_rating.name if a.risk_rating else ""
            atext = f"⚑  {_strip(a.description)}" + (f"  [{risk}]" if risk else "")
            a_vals = ["", "ASSUMPTION", f"    {atext}", "", "", "", "", "", ""]
            for c, v in enumerate(a_vals, 1):
                _cell(ws, row, c, v,
                      font=Font(italic=True, size=9, color="92400E"),
                      fill=_fill(C["l_amber"]), align=AW, border=THIN_B)
            ws.row_dimensions[row].height = 18
            row += 1

        for child in children_map.get(r.id, []):
            _write_result(child, depth + 1)

        if include_activities:
            for act in sorted(r.activities, key=lambda a: a.order):
                total_bgt = sum(bl.amount or 0 for bl in act.budget_lines)
                bg = _fill(C["white"]) if row % 2 == 0 else _fill(C["alt"])
                act_vals = [
                    "", "ACTIVITY",
                    f"{_indent(depth + 1)}↳  {_strip(act.name)}",
                    "",
                    _strip(act.deliverables) or "—",
                    "",
                    _fmt_date(act.start_date),
                    _fmt_date(act.end_date),
                    _fmt_currency(total_bgt, currency),
                ]
                for c, v in enumerate(act_vals, 1):
                    _cell(ws, row, c, v,
                          font=Font(size=9, color=C["dark"]),
                          fill=bg, align=AW, border=THIN_B)
                ws.row_dimensions[row].height = 18
                row += 1

    for top in children_map.get(None, []):
        _write_result(top, 0)

    row += 1
    _write_footer(ws, row, ncols, export_date)
    _set_widths(ws, EXP_WIDTHS)
    ws.freeze_panes = f"A{freeze_row}"


# ---------------------------------------------------------------------------
# SIMPLE FORMAT
# ---------------------------------------------------------------------------
SIMPLE_HEADERS = [
    "Code", "Level", "Name", "Description",
    "Source of Verification", "Baseline", "Targets",
    "Start Date", "End Date", "Budget",
]
SIMPLE_WIDTHS = [8, 12, 40, 44, 28, 12, 26, 12, 12, 16]


def _write_simple_sheet(
    wb: Workbook, *,
    logframe_name: str, org_name: str, project_name: str,
    period: str, currency: str, export_date: str,
    results: list, children_map: dict, codes: dict[int, str],
    level_labels: dict[int, str], all_targets: dict[int, list],
    summary: dict, include_activities: bool, include_summary: bool,
) -> None:
    ws = wb.create_sheet("Logframe")
    ncols = len(SIMPLE_HEADERS)

    row = _write_header(
        ws, 1, ncols,
        logframe_name=logframe_name, org_name=org_name,
        project_name=project_name, period=period,
        currency=currency, export_date=export_date,
    )
    if include_summary:
        row = _write_summary(ws, row, ncols, summary, level_labels)

    _column_headers(ws, row, SIMPLE_HEADERS)
    freeze_row = row + 1
    row += 1

    def _write_result(r, depth: int = 0) -> None:
        nonlocal row
        code = codes.get(r.id, "")
        level = r.level or (depth + 1)
        lbl = level_labels.get(level, f"Level {level}")
        font, fill = _level_style(level)

        inds = sorted(r.indicators, key=lambda i: i.order)
        acts = sorted(r.activities, key=lambda a: a.order)
        total_bgt = sum(bl.amount or 0 for act in acts for bl in act.budget_lines)
        earliest = min((a.start_date for a in acts if a.start_date), default=None)
        latest = max((a.end_date for a in acts if a.end_date), default=None)

        ind_names = "; ".join(_strip(i.name) for i in inds) or "—"
        targets = "; ".join(
            _format_targets(all_targets.get(i.id, []), i.subindicators) for i in inds
        ) or "—"
        mov = "; ".join(_strip(i.source_of_verification) for i in inds if i.source_of_verification) or "—"

        vals = [
            code, lbl.upper(),
            f"{_indent(depth)}{_strip(r.name)}",
            _strip(r.description),
            mov,
            "Required" if any(i.needs_baseline for i in inds) else "—",
            targets,
            _fmt_date(earliest), _fmt_date(latest),
            _fmt_currency(total_bgt, currency),
        ]
        for c, v in enumerate(vals, 1):
            _cell(ws, row, c, v, font=font, fill=fill, align=AW, border=THIN_B)
        ws.row_dimensions[row].height = 20
        row += 1

        for child in children_map.get(r.id, []):
            _write_result(child, depth + 1)

    for top in children_map.get(None, []):
        _write_result(top, 0)

    row += 1
    _write_footer(ws, row, ncols, export_date)
    _set_widths(ws, SIMPLE_WIDTHS)
    ws.freeze_panes = f"A{freeze_row}"


# ---------------------------------------------------------------------------
# EU LOGICAL FRAMEWORK FORMAT
# ---------------------------------------------------------------------------
EU_HEADERS = [
    "Level",
    "Intervention Logic",
    "Objectively Verifiable Indicators (OVI)",
    "Sources of Verification",
    "Assumptions / Risks",
]
EU_WIDTHS = [14, 55, 45, 32, 38]


def _write_eu_sheet(
    wb: Workbook, *,
    logframe_name: str, org_name: str, project_name: str,
    period: str, currency: str, export_date: str,
    results: list, children_map: dict, codes: dict[int, str],
    level_labels: dict[int, str], all_targets: dict[int, list],
    summary: dict, include_summary: bool,
    **_kwargs,
) -> None:
    ws = wb.create_sheet("EU LFA")
    ncols = len(EU_HEADERS)

    row = _write_header(
        ws, 1, ncols,
        logframe_name=logframe_name, org_name=org_name,
        project_name=project_name, period=period,
        currency=currency, export_date=export_date,
    )
    if include_summary:
        row = _write_summary(ws, row, ncols, summary, level_labels)

    _column_headers(ws, row, EU_HEADERS)
    freeze_row = row + 1
    row += 1

    def _write_result(r, depth: int = 0) -> None:
        nonlocal row
        level = r.level or (depth + 1)
        lbl = level_labels.get(level, f"Level {level}")
        font, fill = _level_style(level)

        inds = sorted(r.indicators, key=lambda i: i.order)
        ovi = "\n".join(
            f"• {_strip(i.name)}: {_format_targets(all_targets.get(i.id, []), i.subindicators)}"
            for i in inds
        ) or "—"
        mov = "\n".join(_strip(i.source_of_verification) for i in inds if i.source_of_verification) or "—"
        ass = "\n".join(
            f"• {_strip(a.description)}" + (f"  [{a.risk_rating.name}]" if a.risk_rating else "")
            for a in r.assumptions
        ) or "—"

        vals = [lbl.upper(), _strip(r.name), ovi, mov, ass]
        for c, v in enumerate(vals, 1):
            _cell(ws, row, c, v, font=font, fill=fill, align=AW, border=THIN_B)
        ws.row_dimensions[row].height = max(22, min(80, (len(inds) + 1) * 15))
        row += 1

        for child in children_map.get(r.id, []):
            _write_result(child, depth + 1)

    for top in children_map.get(None, []):
        _write_result(top, 0)

    row += 1
    _write_footer(ws, row, ncols, export_date)
    _set_widths(ws, EU_WIDTHS)
    ws.freeze_panes = f"A{freeze_row}"


# ---------------------------------------------------------------------------
# DFAT FORMAT
# ---------------------------------------------------------------------------
DFAT_HEADERS = [
    "Level", "Result Statement", "Performance Indicator",
    "Means of Verification", "Baseline", "Target",
    "Responsible Party", "Assumptions / Risks",
]
DFAT_WIDTHS = [14, 46, 38, 28, 14, 24, 18, 34]


def _write_dfat_sheet(
    wb: Workbook, *,
    logframe_name: str, org_name: str, project_name: str,
    period: str, currency: str, export_date: str,
    results: list, children_map: dict, codes: dict[int, str],
    level_labels: dict[int, str], all_targets: dict[int, list],
    summary: dict, include_summary: bool,
    **_kwargs,
) -> None:
    ws = wb.create_sheet("DFAT PAF")
    ncols = len(DFAT_HEADERS)

    row = _write_header(
        ws, 1, ncols,
        logframe_name=logframe_name, org_name=org_name,
        project_name=project_name, period=period,
        currency=currency, export_date=export_date,
    )
    if include_summary:
        row = _write_summary(ws, row, ncols, summary, level_labels)

    _column_headers(ws, row, DFAT_HEADERS)
    freeze_row = row + 1
    row += 1

    def _write_result(r, depth: int = 0) -> None:
        nonlocal row
        level = r.level or (depth + 1)
        lbl = level_labels.get(level, f"Level {level}")
        font, fill = _level_style(level)

        inds = sorted(r.indicators, key=lambda i: i.order)
        ind_text = "\n".join(f"• {_strip(i.name)}" for i in inds) or "—"
        mov_text = "\n".join(_strip(i.source_of_verification) for i in inds if i.source_of_verification) or "—"
        t_text = "\n".join(
            _format_targets(all_targets.get(i.id, []), i.subindicators) for i in inds
        ) or "—"
        ass_text = "\n".join(
            f"• {_strip(a.description)}" for a in r.assumptions
        ) or "—"

        vals = [
            lbl.upper(), _strip(r.name),
            ind_text, mov_text,
            "Required" if any(i.needs_baseline for i in inds) else "—",
            t_text, "", ass_text,
        ]
        for c, v in enumerate(vals, 1):
            _cell(ws, row, c, v, font=font, fill=fill, align=AW, border=THIN_B)
        ws.row_dimensions[row].height = max(22, min(80, (len(inds) + 1) * 15))
        row += 1

        for child in children_map.get(r.id, []):
            _write_result(child, depth + 1)

    for top in children_map.get(None, []):
        _write_result(top, 0)

    row += 1
    _write_footer(ws, row, ncols, export_date)
    _set_widths(ws, DFAT_WIDTHS)
    ws.freeze_panes = f"A{freeze_row}"


# ---------------------------------------------------------------------------
# ACTIVITY PLAN SHEET
# ---------------------------------------------------------------------------

def _write_activities_sheet(
    wb: Workbook,
    results: list,
    children_map: dict,
    codes: dict[int, str],
    level_labels: dict[int, str],
    currency: str,
    export_date: str,
) -> None:
    ws = wb.create_sheet("Activity Plan")
    ncols = 8

    _merge(ws, 1, 1, ncols, "ACTIVITY PLAN",
           font=Font(bold=True, size=12, color=C["white"]),
           fill=_fill(C["navy"]), align=AWC)
    ws.row_dimensions[1].height = 28

    headers = ["Code", "Activity", "Under Result", "Start Date", "End Date", "Duration", "Budget", "Deliverables"]
    widths  = [8, 52, 36, 14, 14, 12, 16, 40]
    _column_headers(ws, 2, headers)
    ws.freeze_panes = "A3"
    row = 3

    alt = False

    def _collect(r, depth: int = 0) -> None:
        nonlocal row, alt
        code = codes.get(r.id, "")
        level = r.level or (depth + 1)
        lbl = level_labels.get(level, f"Level {level}")
        acts = sorted(r.activities, key=lambda a: a.order)

        if acts:
            _, grp_fill = _level_style(level)
            grp_font = Font(bold=True, size=9, color=C["white"] if level <= 2 else C["dark"])
            _merge(ws, row, 1, ncols,
                   f"[{lbl.upper()}]  {_strip(r.name)}",
                   font=grp_font, fill=grp_fill, align=AW)
            ws.row_dimensions[row].height = 18
            row += 1

            for i, act in enumerate(acts, 1):
                total_bgt = sum(bl.amount or 0 for bl in act.budget_lines)
                start, end = act.start_date, act.end_date
                duration = ""
                if start and end:
                    days = (end - start).days
                    duration = f"{days // 365}y {(days % 365) // 30}m" if days >= 365 else f"{max(1, days // 30)}m"

                bg = _fill(C["white"]) if alt else _fill(C["alt"])
                alt = not alt
                act_vals = [
                    f"{code}.{i}", _strip(act.name), _strip(r.name),
                    _fmt_date(start), _fmt_date(end), duration,
                    _fmt_currency(total_bgt, currency),
                    _strip(act.deliverables) or "—",
                ]
                for c, v in enumerate(act_vals, 1):
                    _cell(ws, row, c, v, font=Font(size=9), fill=bg, align=AW, border=THIN_B)
                ws.row_dimensions[row].height = 18
                row += 1

        for child in children_map.get(r.id, []):
            _collect(child, depth + 1)

    for top in children_map.get(None, []):
        _collect(top, 0)

    row += 1
    _write_footer(ws, row, ncols, export_date)
    _set_widths(ws, widths)


# ---------------------------------------------------------------------------
# INDICATOR DETAILS SHEET
# ---------------------------------------------------------------------------

def _write_indicators_sheet(
    wb: Workbook,
    results: list,
    codes: dict[int, str],
    level_labels: dict[int, str],
    all_targets: dict[int, list],
    export_date: str,
) -> None:
    ws = wb.create_sheet("Indicator Details")
    ncols = 8

    _merge(ws, 1, 1, ncols, "INDICATOR DETAILS",
           font=Font(bold=True, size=12, color=C["white"]),
           fill=_fill(C["navy"]), align=AWC)
    ws.row_dimensions[1].height = 28

    headers = [
        "Code", "Indicator", "Under Result",
        "Components / Sub-indicators", "Source of Verification",
        "Baseline Required", "Target", "Notes",
    ]
    widths = [8, 50, 36, 32, 30, 14, 30, 18]
    _column_headers(ws, 2, headers)
    ws.freeze_panes = "A3"
    row = 3

    alt = False
    ind_seq = 0

    for r in results:
        if not r.indicators:
            continue
        code = codes.get(r.id, "")
        level = r.level
        lbl = level_labels.get(level, f"Level {level}") if level else ""

        for ind in sorted(r.indicators, key=lambda i: i.order):
            ind_seq += 1
            targets = all_targets.get(ind.id, [])
            subs = ind.subindicators
            components = "; ".join(s.name for s in subs) if subs else "—"
            t_text = _format_targets(targets, subs)

            bg = _fill(C["white"]) if alt else _fill(C["alt"])
            alt = not alt
            vals = [
                f"{code}.I{ind_seq}",
                _strip(ind.name),
                f"[{lbl.upper()}]  {_strip(r.name)}" if lbl else _strip(r.name),
                components,
                _strip(ind.source_of_verification) or "—",
                "Yes" if ind.needs_baseline else "No",
                t_text,
                "",
            ]
            for c, v in enumerate(vals, 1):
                _cell(ws, row, c, v, font=Font(size=9), fill=bg, align=AW, border=THIN_B)
            ws.row_dimensions[row].height = max(18, min(60, len(t_text.split("\n")) * 14))
            row += 1

    row += 1
    _write_footer(ws, row, ncols, export_date)
    _set_widths(ws, widths)


# ---------------------------------------------------------------------------
# BUDGET SUMMARY SHEET
# ---------------------------------------------------------------------------

def _write_budget_sheet(
    wb: Workbook,
    results: list,
    children_map: dict,
    codes: dict[int, str],
    level_labels: dict[int, str],
    currency: str,
    export_date: str,
    summary: dict,
) -> None:
    ws = wb.create_sheet("Budget Summary")
    ncols = 6

    _merge(ws, 1, 1, ncols, "BUDGET SUMMARY",
           font=Font(bold=True, size=12, color=C["white"]),
           fill=_fill(C["navy"]), align=AWC)
    ws.row_dimensions[1].height = 28

    _merge(ws, 2, 1, ncols,
           f"Total Project Budget: {summary['total_budget_fmt']}",
           font=Font(bold=True, size=11, color=C["navy"]),
           fill=_fill(C["l_blue"]), align=AWC)
    _row_border(ws, 2, 1, ncols, THIN_B)
    ws.row_dimensions[2].height = 24

    headers = ["Code", "Activity", "Under Result", "Budget Line", "Amount", "Activity Total"]
    widths  = [8, 52, 36, 26, 16, 16]
    _column_headers(ws, 4, headers)
    ws.freeze_panes = "A5"
    row = 5

    alt = False
    act_seq = 0

    def _collect(r, depth: int = 0) -> None:
        nonlocal row, alt, act_seq
        code = codes.get(r.id, "")
        level = r.level or (depth + 1)
        lbl = level_labels.get(level, f"Level {level}")

        for act in sorted(r.activities, key=lambda a: a.order):
            act_seq += 1
            total_bgt = sum(bl.amount or 0 for bl in act.budget_lines)
            act_code = f"{code}.A{act_seq}"

            if act.budget_lines:
                for i, bl in enumerate(act.budget_lines):
                    bg = _fill(C["white"]) if alt else _fill(C["alt"])
                    vals = [
                        act_code if i == 0 else "",
                        _strip(act.name) if i == 0 else "",
                        f"[{lbl.upper()}]  {_strip(r.name)}" if i == 0 else "",
                        _strip(bl.name) or f"Budget line {i + 1}",
                        _fmt_currency(bl.amount or 0, currency),
                        _fmt_currency(total_bgt, currency) if i == 0 else "",
                    ]
                    for c, v in enumerate(vals, 1):
                        _cell(ws, row, c, v, font=Font(size=9), fill=bg, align=AW, border=THIN_B)
                    ws.row_dimensions[row].height = 18
                    row += 1
            else:
                bg = _fill(C["white"]) if alt else _fill(C["alt"])
                vals = [
                    act_code, _strip(act.name),
                    f"[{lbl.upper()}]  {_strip(r.name)}",
                    "—", "—",
                    _fmt_currency(0, currency),
                ]
                for c, v in enumerate(vals, 1):
                    _cell(ws, row, c, v, font=Font(size=9), fill=bg, align=AW, border=THIN_B)
                ws.row_dimensions[row].height = 18
                row += 1

            alt = not alt

        for child in children_map.get(r.id, []):
            _collect(child, depth + 1)

    for top in children_map.get(None, []):
        _collect(top, 0)

    # Grand total row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value="TOTAL")
    c.font = Font(bold=True, size=10, color=C["white"])
    c.fill = _fill(C["navy"])
    c.alignment = AWR
    c.border = THIN_B
    _cell(ws, row, 6, summary["total_budget_fmt"],
          font=Font(bold=True, size=10, color=C["white"]),
          fill=_fill(C["navy"]), align=AWR, border=THIN_B)
    ws.row_dimensions[row].height = 22
    row += 1

    row += 1
    _write_footer(ws, row, ncols, export_date)
    _set_widths(ws, widths)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def build_logframe_excel(
    *,
    logframe_name: str,
    org_name: str = "",
    project_name: str = "",
    currency: str = "USD",
    start_year: int | None = None,
    end_year: int | None = None,
    results: list,
    all_targets: dict[int, list],
    level_labels: dict[int, str],
    style: str = "donor",
    include_activities: bool = True,
    include_indicators: bool = True,
    include_budget: bool = True,
    include_summary: bool = True,
) -> io.BytesIO:
    """Build and return an Excel workbook as a BytesIO buffer."""

    export_date = datetime.now().strftime("%d %B %Y")
    period = _fmt_period(start_year, end_year)

    children_map: dict[int | None, list] = defaultdict(list)
    for r in results:
        children_map[r.parent_id].append(r)
    for lst in children_map.values():
        lst.sort(key=lambda x: x.order)

    codes = _generate_codes(results)
    summary = _compute_summary(results, currency)

    common = dict(
        logframe_name=logframe_name,
        org_name=org_name,
        project_name=project_name,
        period=period,
        currency=currency,
        export_date=export_date,
        results=results,
        children_map=children_map,
        codes=codes,
        level_labels=level_labels,
        all_targets=all_targets,
        summary=summary,
        include_activities=include_activities,
        include_summary=include_summary,
    )

    wb = Workbook()
    wb.remove(wb.active)

    STYLE_MAP = {
        "simple":   _write_simple_sheet,
        "expanded": _write_expanded_sheet,
        "eu":       _write_eu_sheet,
        "dfat":     _write_dfat_sheet,
    }
    writer = STYLE_MAP.get(style, _write_donor_sheet)
    writer(wb, **common)

    if include_activities:
        _write_activities_sheet(wb, results, children_map, codes, level_labels, currency, export_date)
    if include_indicators:
        _write_indicators_sheet(wb, results, codes, level_labels, all_targets, export_date)
    if include_budget and include_activities:
        _write_budget_sheet(wb, results, children_map, codes, level_labels, currency, export_date, summary)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
